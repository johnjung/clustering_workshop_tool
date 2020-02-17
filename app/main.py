from datetime import datetime
from flask import Flask, Response, render_template, request
from io import BytesIO, StringIO
from matplotlib import pyplot as plt
from networkx.drawing.nx_agraph import graphviz_layout
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from scipy.cluster.hierarchy import dendrogram, linkage
from werkzeug.wsgi import FileWrapper

import csv
import matplotlib
import networkx
import numpy as np
import random
import requests
import string
import sys

app = Flask(__name__)

# TODO
# see: https://docs.google.com/spreadsheets/d/1IeN0K6LpULcYiqetBl4koNZA1PkKx-NftKgBr0PNxzc/edit#gid=1467402753
# the labels are a bit long for the graph. Is there anything I can do to help that? 

# set matplotlib to use a non-interactive backend- i.e., don't try to create windows on the server.
matplotlib.use('agg')

def get_google_sheets_csv_url(u):
    """
    e.g. https://docs.google.com/spreadsheets/d/1s2Q0wC1TheZIahkw_Ly876pZUNCUZv82bVMu2A9f774/edit?usp=sharing
    to-  https://docs.google.com/spreadsheets/d/1s2Q0wC1TheZIahkw_Ly876pZUNCUZv82bVMu2A9f774/gviz/tq?tqx=out:csv
    """
    pieces = u.split('/')[:6]
    pieces.append('export?format=csv')
    return '/'.join(pieces)

def get_hex(value, max_value):
    """
    0 -> #ffffff
    3 -> #666666
    """
    r = float(value) / float(max_value)
    h =  hex(int(255 - (128.0 * r))).upper()[2:]
    return 'FF{0}{0}'.format(h)

def get_random_filename():
    return ''.join(random.choice(string.ascii_lowercase) for i in range(8))

def get_headers_data_maxvalue(csvstring):
    # catch errors: the url might not be public. 
    if '<!DOCTYPE html>' in csvstring:
        return render_template('error.html')

    f = StringIO(csvstring)
    reader = csv.reader(f, delimiter=",")

    headings_1 = next(reader)[1:]
    headings_2 = []
    data = []

    # input data...
    # cells should be an empty string ('') if data was missing or not an
    # integer, or an integer.
    for row in reader:
        headings_2.append(row[0])
        tmp_row = []
        for cell in row[1:]:
            if cell == '':
                tmp_row.append(cell)
            else:
                try:
                    tmp_row.append(int(cell))
                except ValueError:
                    tmp_row.append('')
        data.append(tmp_row)

    # labels must be in the right orders, in the right positions.  
    try:
        assert headings_1 == headings_2 
    except:
        return ' '.join(headings_1) + ' !!! ' + ' '.join(headings_2)

    # get the maximum value from the input data.
    max_value = 0
    for y in range(len(data)):
        for x in range(len(data)):
            try:
                if data[y][x] > max_value:
                    max_value = data[y][x]
            except TypeError:
                pass

    # "fill in" the input data. 
    # set the diagonal to the max value.
    # if cells are empty, set them to zero.
    # if a corresponding cell has a value, copy it's value to this cell.
    # if cells have different values, set them to zero.
    y = 0
    while y < len(data):
        x = 0
        while x <= y:
            if x == y:
                data[y][x] = max_value

            if data[y][x] == '' and data[x][y] == '':
                data[y][x] = 0
                data[x][y] = 0
            elif data[y][x] == '':
                data[y][x] = data[x][y]
            elif data[x][y] == '':
                data[x][y] = data[y][x]
            elif data[y][x] != data[x][y]:
                data[y][x] = 0
                data[x][y] = 0
            x += 1
        y += 1
    return headings_1, data, max_value

@app.route("/")
def form():
    """
    Display a webform to the user.
    """
    return render_template('form.html')

@app.route("/cluster", methods=["POST"])
def cluster():
    # clear the current figure- so repeated requests to the service will render correctly.
    plt.clf()

    if request.form['url'] == '':
        return render_template('error.html')

    url = get_google_sheets_csv_url(request.form["url"])
    headers, data, max_value = get_headers_data_maxvalue(requests.get(url).text)

    if request.form['output'] in ('matrix', 'dendrogram'):
        l = linkage(data, request.form['linkage'])
        d = dendrogram(
            l,  
            above_threshold_color='k', 
            color_threshold=0, 
            labels=headers,
            orientation='left'
        ) 

    if request.form['output'] == 'matrix':
        # sort the matrix based on dendrogram order.
        data_out = []
        for y in range(len(data)):
            row_out = []
            for x in range(len(data)):
                row_out.append(data[d['leaves'][y]][d['leaves'][x]])
            data_out.append(row_out)
    
        # make an xlsx spreadsheet here to download. 
        wb = Workbook()
        ws = wb.active
        for y in range(len(data_out)):            
            for x in range(len(data_out)):            
                cell = ws.cell(row=y+2, column=x+2, value=data_out[y][x])
                color = get_hex(data_out[y][x], max_value)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
        for i in range(len(data_out)):
            ws.cell(row=1, column=i+2, value=headers[d['leaves'][i]])
            ws.cell(row=i+2, column=1, value=headers[d['leaves'][i]])
    
        f = BytesIO()
        wb.save(f)
        f.seek(0)
        w = FileWrapper(f)
        try:
            return Response(w, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", direct_passthrough=True)
        except Exception as e:
            return str(e)
    elif request.form['output'] == 'dendrogram':
        plt.subplots_adjust(right=0.66)
        f = BytesIO()
        plt.savefig(f, format='svg')
        f.seek(0)
        w = FileWrapper(f)
        try:
            r = Response(w, mimetype="image/svg+xml", direct_passthrough=True)
            r.headers['Content-Disposition'] = 'attachment; filename="{}.svg"'.format(get_random_filename())
            return r
        except Exception as e:
            return str(e)
    elif request.form['output'] == 'graph':
        g = networkx.Graph()
        for a in range(len(headers)):
            g.add_node(headers[a])
            for b in range(a):
                if data[a][b] >= int(request.form['cutoff']):
                    g.add_edge(headers[a], headers[b])
                    g.add_edge(headers[b], headers[a])
    
        pos = graphviz_layout(g, prog='neato')
    
        networkx.draw_networkx_nodes(g, pos, node_color='#ffffff')
        networkx.draw_networkx_labels(g, pos)
        networkx.draw_networkx_edges(g, pos)
    
        plt.subplots_adjust(top=1, bottom=0, right=1, left=0, hspace=0, wspace=0)
        f = BytesIO()
        plt.savefig(f, bbox_inches='tight', edgecolor='w', facecolor='w', format='svg', pad_inches=0)
        f.seek(0)
        w = FileWrapper(f)
        try:
            r = Response(w, mimetype="image/svg+xml", direct_passthrough=True)
            r.headers['Content-Disposition'] = 'attachment; filename="{}.svg"'.format(get_random_filename())
            return r
        except Exception as e:
            return str(e)
    else:
        raise NotImplementedError

'''
if __name__ == "__main__":
    # debug only
    app.run(host='0.0.0.0', debug=True, port=80)
'''
